"""Excel/CSV -> DuckDB. The workbook's numbers stay exact and queryable; only a
short schema summary goes into the LightRAG graph (see build_summary)."""

import datetime
import logging
import os
import re
import uuid
from io import BytesIO

import duckdb
import openpyxl

from app.core.config import get_settings

logger = logging.getLogger("app.parsers.spreadsheet")
_settings = get_settings()

SPREADSHEET_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}

# First line of build_summary(). The entity extractor matches on it to leave
# spreadsheet documents alone -- their structure is written into the graph
# deterministically, so extracting it again from prose can only add noise.
SUMMARY_HEADER = "Spreadsheet workbook:"

_CELL_REF = re.compile(r"\$?([A-Z]{1,3})\$?\d+")
_CURRENCY_FORMAT = re.compile(r"[$€£¥₹]|\b(usd|eur|gbp|inr)\b", re.IGNORECASE)

# One connection per user, keyed by that user's storage directory. A user's
# tables live in their own database file, so a generated SELECT cannot name
# another user's table even if the SQL validator were bypassed entirely --
# isolation by construction rather than by a filter on a query an LLM wrote.
_connections: dict[str, duckdb.DuckDBPyConnection] = {}


def close_connections() -> None:
    """Release every open DuckDB file (shutdown, and between tests)."""
    for connection in _connections.values():
        connection.close()
    _connections.clear()


def _evict_idle_connections() -> None:
    """Keep at most `max_active_workspaces` DuckDB files open.

    Each open connection holds a file handle and a WAL; a beta with more users
    than active ones should not hold every one of them open forever. Dropping
    the oldest is safe: the next call simply reopens the file.

    ponytail: insertion-order eviction, not true LRU. Swap in an OrderedDict
    move_to_end if a user's connection ever gets dropped mid-session often
    enough to matter.
    """
    while len(_connections) > _settings.max_active_workspaces:
        _, connection = _connections.popitem()
        connection.close()


def get_connection() -> duckdb.DuckDBPyConnection:
    """The current user's tabular store. One file per user, one connection.

    External access is off for the life of the connection: an LLM-generated
    SELECT can otherwise reach the local filesystem through read_csv/read_text.
    Ingestion sniffs CSVs on a separate scratch connection instead.

    ponytail: one connection per user, serialized by the event loop. Hand out
    `con.cursor()` per request if concurrent queries ever matter.
    """
    from app.core import auth

    directory = auth.user_dir()
    if directory not in _connections:
        os.makedirs(directory, exist_ok=True)
        _connection = duckdb.connect(
            os.path.join(directory, "spreadsheets.duckdb"),
            config={"enable_external_access": False},
        )
        # Rebrand migration: the metadata table used to be `_crag_columns`, and
        # it lives in a file that outlives the rename. Idempotent -- a fresh DB
        # has neither table and the CREATE below makes the new one.
        if _connection.execute(
            "SELECT count(*) FROM duckdb_tables() WHERE table_name = '_crag_columns'"
        ).fetchone()[0]:
            _connection.execute("ALTER TABLE _crag_columns RENAME TO _node_rels_columns")
        _connection.execute(
            """CREATE TABLE IF NOT EXISTS _node_rels_columns (
                   table_name TEXT, column_name TEXT, data_type TEXT,
                   semantic TEXT, formula TEXT, derived_from TEXT,
                   workbook TEXT, worksheet TEXT, added_later BOOLEAN DEFAULT FALSE)"""
        )
        _connections[directory] = _connection
        _evict_idle_connections()
    return _connections[directory]


def _identifier(name: str) -> str:
    cleaned = re.sub(r"\W+", "_", str(name)).strip("_").lower()
    return cleaned or "col"


def _quote(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _semantic_type(values: list, number_format: str) -> tuple[str, str]:
    """(semantic label, DuckDB type). Excel keeps dates as serial numbers and
    percentages as plain floats, so the cell's number format -- not the raw
    value -- is what tells them apart."""
    sample = [v for v in values if v not in (None, "")]
    if not sample:
        return "text", "VARCHAR"

    if all(isinstance(v, bool) for v in sample):
        return "boolean", "BOOLEAN"
    if all(isinstance(v, datetime.datetime) for v in sample):
        return "date", "TIMESTAMP"
    if all(isinstance(v, datetime.date) for v in sample):
        return "date", "DATE"
    if all(isinstance(v, (int, float)) for v in sample):
        if "%" in number_format:
            return "percentage", "DOUBLE"
        if _CURRENCY_FORMAT.search(number_format):
            return "currency", "DOUBLE"
        if all(isinstance(v, int) for v in sample):
            return "numeric", "BIGINT"
        return "numeric", "DOUBLE"

    distinct = len({str(v) for v in sample})
    if distinct <= max(2, len(sample) // 4):
        return "categorical", "VARCHAR"
    return "text", "VARCHAR"


def _header_letters(ws) -> dict[str, str]:
    """Column letter -> header text, for resolving formula references."""
    return {
        cell.column_letter: str(cell.value).strip()
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
        if cell.value is not None
    }


def _formula_lineage(formula: str, letters: dict[str, str], own: str) -> list[str]:
    referenced = {
        letters[m.group(1)]
        for m in _CELL_REF.finditer(formula)
        if m.group(1) in letters
    }
    return sorted(referenced - {own})


def _scan_cells(ws, limit: int = 20):
    """Formulas and number formats from the first data rows -- both are
    per-column properties in practice, so a sample is enough."""
    formulas: dict[str, str] = {}
    formats: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, limit + 1)):
        for cell in row:
            if cell.value is None:
                continue
            formats.setdefault(cell.column_letter, cell.number_format)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.setdefault(cell.column_letter, cell.value)
    return formulas, formats


def _read_sheet(ws_values, ws_formulas) -> tuple[list[dict], list[tuple]] | None:
    """(column specs, rows) for one worksheet, or None if it holds no table."""
    rows = list(ws_values.iter_rows(values_only=True))
    if len(rows) < 2:
        return None

    headers = [
        str(h).strip() if h is not None else f"column_{i + 1}"
        for i, h in enumerate(rows[0])
    ]
    if not headers:
        return None

    body = [r for r in rows[1:] if any(c is not None and c != "" for c in r)]
    if not body:
        return None

    letters = _header_letters(ws_formulas)
    formulas, formats = _scan_cells(ws_formulas)

    columns = []
    for index, header in enumerate(headers):
        letter = openpyxl.utils.get_column_letter(index + 1)
        values = [row[index] if index < len(row) else None for row in body]
        semantic, data_type = _semantic_type(values, formats.get(letter, ""))
        formula = formulas.get(letter)
        columns.append(
            {
                "name": _identifier(header),
                "header": header,
                "data_type": data_type,
                "semantic": semantic,
                "formula": formula,
                "derived_from": _formula_lineage(formula, letters, header) if formula else [],
            }
        )

    width = len(columns)
    padded = [tuple(list(r[:width]) + [None] * (width - len(r))) for r in body]
    return columns, padded


def _create_table(con, table: str, columns: list[dict], rows: list[tuple]) -> None:
    cols_sql = ", ".join(f"{_quote(c['name'])} {c['data_type']}" for c in columns)
    con.execute(f"DROP TABLE IF EXISTS {_quote(table)}")
    con.execute(f"CREATE TABLE {_quote(table)} ({cols_sql})")
    if rows:
        placeholders = ", ".join("?" * len(columns))
        con.executemany(f"INSERT INTO {_quote(table)} VALUES ({placeholders})", rows)


def record_columns(con, table, columns, workbook, worksheet, added_later=False) -> None:
    con.execute(
        "DELETE FROM _node_rels_columns WHERE table_name = ? AND column_name IN "
        f"({', '.join('?' * len(columns))})",
        [table, *[c["name"] for c in columns]],
    )
    con.executemany(
        "INSERT INTO _node_rels_columns VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [
            (
                table,
                c["name"],
                c["data_type"],
                c["semantic"],
                c["formula"],
                ", ".join(c["derived_from"]) or None,
                workbook,
                worksheet,
                added_later,
            )
            for c in columns
        ],
    )


def _load_csv(con, data: bytes, table: str, file_name: str, stem: str) -> dict:
    """DuckDB's CSV sniffer does the type inference, but it runs on a throwaway
    in-memory connection: the real store has file access switched off (see
    get_connection), which is what stops a generated SELECT reading the disk.

    ponytail: whole CSV materialized in Python on the way across. Fine at
    spreadsheet scale; ATTACH the temp database instead if that stops being true.
    """
    path = os.path.join(_settings.storage_dir, f"upload_{uuid.uuid4().hex[:8]}.csv")
    with open(path, "wb") as f:
        f.write(data)
    try:
        scratch = duckdb.connect()
        scratch.execute("CREATE TABLE t AS SELECT * FROM read_csv_auto(?)", [path])
        schema = scratch.execute("DESCRIBE t").fetchall()
        rows = scratch.execute("SELECT * FROM t").fetchall()
        scratch.close()
    finally:
        os.remove(path)

    columns = [
        {
            "name": _identifier(row[0]),
            "header": row[0],
            "data_type": row[1],
            "semantic": "text" if row[1].upper().startswith("VARCHAR") else "numeric",
            "formula": None,
            "derived_from": [],
        }
        for row in schema
    ]
    _create_table(con, table, columns, rows)
    record_columns(con, table, columns, file_name, stem)
    return {"worksheet": stem, "table": table, "rows": len(rows), "columns": columns}


def load_spreadsheet(data: bytes, file_name: str) -> dict:
    """Load every worksheet into its own DuckDB table. Returns the workbook
    structure, which is also the input to build_summary()."""
    con = get_connection()
    # Re-uploading the same workbook replaces it. Without this the ingest
    # dedup keeps one manifest row while DuckDB accumulates a second copy of
    # every table, and the query layer sees both.
    drop_workbook_tables(file_name)
    workbook_id = uuid.uuid4().hex[:8]
    stem = _identifier(os.path.splitext(file_name)[0])
    ext = os.path.splitext(file_name)[1].lower()
    sheets: list[dict] = []

    if ext == ".csv":
        table = f"workbook_{workbook_id}__{stem}"
        sheets.append(_load_csv(con, data, table, file_name, stem))
    else:
        values_wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        formula_wb = openpyxl.load_workbook(BytesIO(data), data_only=False)
        for sheet_name in values_wb.sheetnames:
            parsed = _read_sheet(values_wb[sheet_name], formula_wb[sheet_name])
            if parsed is None:
                logger.info("Skipping empty worksheet %r in %s", sheet_name, file_name)
                continue
            columns, rows = parsed
            table = f"workbook_{workbook_id}__{_identifier(sheet_name)}"
            _create_table(con, table, columns, rows)
            record_columns(con, table, columns, file_name, sheet_name)
            sheets.append(
                {
                    "worksheet": sheet_name,
                    "table": table,
                    "rows": len(rows),
                    "columns": columns,
                }
            )

    if not sheets:
        raise ValueError(f"No tabular data found in {file_name}.")

    return {"workbook_id": workbook_id, "file_name": file_name, "sheets": sheets}


def build_summary(workbook: dict) -> str:
    """The document body for a spreadsheet: a short, factual index entry.

    Deliberately thin. The workbook's real structure -- worksheets, columns,
    types, formula lineage -- is written node by node into the knowledge graph
    by `tabular_graph`, and each of those nodes carries its own retrieval card.
    Repeating it here as prose only handed the entity extractor words like
    VARCHAR, BIGINT and `categorical` to turn into graph nodes, which is what
    the graph filled up with. `SUMMARY_HEADER` is how the extractor recognises
    this text and skips it.
    """
    lines = [f"{SUMMARY_HEADER} {workbook['file_name']}"]
    for sheet in workbook["sheets"]:
        lines.append(
            f"Worksheet {sheet['worksheet']!r}: {sheet['rows']} rows, "
            f"{len(sheet['columns'])} columns, queryable as the table "
            f"{sheet['table']} in the nodeRels spreadsheet store."
        )
    return "\n".join(lines)


def drop_workbook_tables(file_name: str) -> None:
    """Called when the document is deleted, so a removed workbook can't keep
    answering queries out of the tabular store."""
    con = get_connection()
    tables = [
        row[0]
        for row in con.execute(
            "SELECT DISTINCT table_name FROM _node_rels_columns WHERE workbook = ?",
            [file_name],
        ).fetchall()
    ]
    for table in tables:
        con.execute(f"DROP TABLE IF EXISTS {_quote(table)}")
    con.execute("DELETE FROM _node_rels_columns WHERE workbook = ?", [file_name])
